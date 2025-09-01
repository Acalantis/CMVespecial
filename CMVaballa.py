import pandas as pd
import os
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import streamlit as st
import zipfile

# Função para buscar preço baseado no SKU (correspondência parcial e sem distinção entre maiúsculas e minúsculas)
def buscar_preco(sku2, df_base, coluna_sku_base, coluna_preco_base):
    sku2 = str(sku2).strip().lower()  # Normaliza para minúsculas e remove espaços extras
    for _, row in df_base.iterrows():
        sku1 = str(row[coluna_sku_base]).strip().lower()  # Normaliza para minúsculas e remove espaços extras
        if sku1 in sku2 or sku2 in sku1:  # Permite correspondência parcial
            return row[coluna_preco_base]  # Retorna o preço correspondente
    return None

# Função para atualizar planilhas
def atualizar_planilha(planilha, planilha_base, coluna_sku_base, coluna_preco_base, coluna_sku_planilha, novo_nome_coluna):
    planilha[novo_nome_coluna] = planilha[coluna_sku_planilha].apply(
        lambda sku: buscar_preco(sku, planilha_base, coluna_sku_base, coluna_preco_base)
    )
    planilha[novo_nome_coluna].fillna("SKU não encontrado", inplace=True)
    return planilha

# Interface do Streamlit
st.title("Atualizador de CMV")

# Upload do arquivo base (CMV Thiago)
arquivo_base = st.file_uploader("Faça upload da planilha CMV (Base de dados)", type=["xlsx"])

# Upload de múltiplos arquivos para atualização
arquivos_para_atualizar = st.file_uploader("Faça upload das planilhas para atualizar", type=["xlsx"], accept_multiple_files=True)

# Botão para processar as planilhas
if st.button("Gerar CMV Atualizado"):
    if not arquivo_base or not arquivos_para_atualizar:
        st.error("Por favor, envie a planilha CMV e as planilhas a serem atualizadas.")
    else:
        coluna_sku_base = "SKU"
        coluna_preco_base = "Preço"
        coluna_sku_planilha = "Número de referência SKU"
        novo_nome_coluna = "CMV"

        # Ler a planilha base
        planilha_base = pd.read_excel(arquivo_base)

        arquivos_atualizados = []

        # Barra de progresso
        progresso = st.progress(0)
        total_arquivos = len(arquivos_para_atualizar)

        # Processar arquivos
        for idx, arquivo in enumerate(arquivos_para_atualizar):
            # Ler cada planilha
            planilha = pd.read_excel(arquivo)

            # Atualizar a planilha
            planilha_atualizada = atualizar_planilha(
                planilha, planilha_base, coluna_sku_base, coluna_preco_base, coluna_sku_planilha, novo_nome_coluna
            )

            # Aplicar estilo na coluna CMV
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                planilha_atualizada.to_excel(writer, index=False, sheet_name="Atualizado")
                workbook = writer.book
                sheet = workbook.active

                # Pintar a coluna CMV de amarelo
                col_idx = list(planilha_atualizada.columns).index(novo_nome_coluna) + 1
                amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                for row in range(2, sheet.max_row + 1):  # Ignorar o cabeçalho
                    cell = sheet.cell(row=row, column=col_idx)
                    cell.fill = amarelo

            output.seek(0)
            arquivos_atualizados.append((arquivo.name.replace(".xlsx", " atualizada.xlsx"), output))

            # Atualizar barra de progresso
            progresso.progress((idx + 1) / total_arquivos)

        # Compactar todos os arquivos em um único ZIP
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zip_file:
            for nome, arquivo in arquivos_atualizados:
                zip_file.writestr(nome, arquivo.getvalue())
        zip_buffer.seek(0)

        # Permitir download do arquivo ZIP
        st.download_button(
            label="Baixar todas as planilhas atualizadas",
            data=zip_buffer,
            file_name="planilhas_atualizadas.zip",
            mime="application/zip",
        )

        st.success("Planilhas atualizadas com sucesso!")